import cv2 as cv
import openpyxl

template="template.jpg"
excel_file="certi_names.xlsx"
output_path=""

font_size=1
font_color=(227,172,30)
thickness=2

y_coordinate=-20
x_coordinate=6

workbook=openpyxl.load_workbook(excel_file)
working_sheet=workbook.active

for i in range(2,3):
    name=(working_sheet.cell(row=i,column=1)).value
    gender=(working_sheet.cell(row=i,column=2)).value
    certi_no=(working_sheet.cell(row=i,column=3)).value
    mentor_name=(working_sheet.cell(row=i,column=4)).value
    project_manager=(working_sheet.cell(row=i,column=5)).value
    if gender.lower()=="m":
        gender="his"
    else:
        gender="her"

    txt1="for satisfactorily completing "+gender+" python Internship"
    txt2=" during ( Nov 2020 - Dec 2020 ) at BestEnlist."

    img=cv.imread(template)
    font=cv.FONT_HERSHEY_SIMPLEX
    text_size=cv.getTextSize(name,font,font_size,cv.LINE_AA)[0]

    text_x = (img.shape[1] - text_size[0]) / 2 + x_coordinate
    text_y = (img.shape[0] + text_size[1]) / 2 - y_coordinate
    text_x = int(text_x)
    text_y = int(text_y)
    cv.putText(img, name, (text_x ,text_y ), font, font_size, font_color,thickness, cv.LINE_AA)
    cv.putText(img, txt1, (text_x-75 ,text_y+25 ), font, 0.45, (0,0,0),thickness=1)
    cv.putText(img, txt2, (text_x-75 ,text_y+45 ), font, 0.45, (0,0,0),thickness=1)
    cv.putText(img,"Certificate number : "+str(certi_no),(text_x-180 ,text_y+80 ),font, 0.4, (100,100,100),thickness=1)
    cv.putText(img,mentor_name,(text_x-120 ,text_y+124 ),font, 0.5, (130,130,130),thickness=1)
    cv.putText(img,project_manager,(text_x+160 ,text_y+124 ),font, 0.5, (130,130,130),thickness=1)

    certi_path = output_path+'certi_'+name+'.png'

    cv.imwrite(certi_path,img)
